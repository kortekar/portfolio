import qrcode
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from tkinter import Tk
from tkinter.filedialog import askopenfilename

def main():
    Tk().withdraw()
    excel_file = askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file:
        print("No file selected.")
        return

    sheet_name = 'Sheet1'

    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    try:
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    qr_column = len(df.columns) + 1
    ws.cell(row=1, column=qr_column, value="QR Code")

    for index, row in df.iterrows():
        qr_text = f"Student ID: {row.get('Student ID', 'N/A')}\n" \
                  f"Name: {row.get('Name', 'N/A')}\n" \
                  f"Grade: {row.get('Grade', 'N/A')}\n" \
                  f"Section: {row.get('Section', 'N/A')}\n" \
                  f"Date of Birth: {row.get('Date of Birth', 'N/A')}\n" \
                  f"Contact Number: {row.get('Contact Number', 'N/A')}\n" \
                  f"Address: {row.get('Address', 'N/A')}"
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_text)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        temp_img_path = f"temp_{row.get('Student ID', 'unknown')}_qr_code.png"
        img.save(temp_img_path)

        img = Image(temp_img_path)
        img.anchor = ws.cell(row=index + 2, column=qr_column).coordinate
        ws.add_image(img)

    wb.save(excel_file)
    print("QR codes generated and added to the Excel sheet successfully.")

if __name__ == "__main__":
    main()
